#!/usr/bin/env python3
"""Import Melipayamak approved SMS patterns from a panel export spreadsheet.

expects columns roughly matching ملی‌پیامک exports:
  کد متن / bodyId → شناسهٔ پترن (bodyId)
  عنوان → عنوان پترن
  متن → متن الگو با متغیرها

Output (default): metadata/melipayamak_patterns.json
Compatible با خروجی خلاصهٔ sync_melipayamak_patterns.py؛ ردیف‌ها شامل فیلدهای اضافی برای UI/فرایندها هستند.

Usage:
  python scripts/import_melipayamak_patterns_xlsx.py path/to/لیست\\ الگوها\\ پیامکی.xlsx
  python scripts/import_melipayamak_patterns_xlsx.py --out metadata/custom.json FILE.xlsx

Requires: openpyxl (در requirements پروژه)
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DEFAULT_OUT = ROOT / "metadata" / "melipayamak_patterns.json"

# تشخیص ستون از سرصفحهٔ فارسی/انگلیسی
_BODY_KEYS = ("کد متن", "کدمتن", "bodyid", "body id", "body_id", "کد الگو", "کدالگو")
_TITLE_KEYS = ("عنوان", "title", "نام الگو")
_TEMPLATE_KEYS = ("متن الگو", "متن پیام", "متن", "template", "بدنه", "متن دارای متغیر")

# یک regex برای حفظ ترتیبٔ ظاهر شدن در متن (مهم برای {0};{1} ملی‌پیامک)
_PLACEHOLDER_RE = re.compile(
    r"%\s*([^%\n\r]+?)\s*%|\{([^}\n\r]+)}|#\s*([^#\n\r]+?)\s*#",
)


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
    return str(v).strip()


def _normalize_header(h: str) -> str:
    t = _cell_str(h).replace("\u200c", "").replace("\xa0", " ")
    return " ".join(t.split())


def _header_matches(header: str, keywords: tuple[str, ...]) -> bool:
    hn = _normalize_header(header).lower()
    for k in keywords:
        kl = k.lower()
        if kl in hn or k in header:
            return True
    return False


def _pick_columns(headers: list[str]) -> tuple[int | None, int | None, int | None]:
    bi = ti = te = None
    for idx, raw in enumerate(headers):
        h = _normalize_header(raw)
        if not h:
            continue
        if bi is None and _header_matches(h, _BODY_KEYS):
            bi = idx
        elif ti is None and _header_matches(h, _TITLE_KEYS):
            ti = idx
        elif te is None and _header_matches(h, _TEMPLATE_KEYS):
            te = idx
    return bi, ti, te


def _parse_body_id(raw: Any) -> int | None:
    s = _cell_str(raw).translate(str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789"))
    s = s.replace(",", "").strip()
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _extract_variables_ordered(template: str) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for m in _PLACEHOLDER_RE.finditer(template or ""):
        raw = next((g for g in m.groups() if g is not None), "")
        name = _normalize_header(raw)
        if name and name not in seen:
            seen.add(name)
            ordered.append(name)
    return ordered


def _first_data_row(ws: Any) -> tuple[list[str], int]:
    """Returns header cells (strings) and 1-based row index of header."""
    import openpyxl  # pylint: disable=import-outside-toplevel

    max_scan = min(ws.max_row or 1, 25)
    best: tuple[list[str], int] | None = None
    for r in range(1, max_scan + 1):
        row_vals = [_cell_str(c.value) for c in ws[r]]
        if not any(row_vals):
            continue
        bi, ti, te = _pick_columns(row_vals)
        if bi is not None and te is not None:
            return row_vals, r
        if bi is not None or te is not None:
            best = (row_vals, r)
    if best:
        return best
    row_vals = [_cell_str(c.value) for c in ws[1]]
    return row_vals, 1


def import_xlsx(path: Path) -> dict[str, Any]:
    import openpyxl  # pylint: disable=import-outside-toplevel

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        headers, header_row = _first_data_row(ws)
        bi, ti, te = _pick_columns(headers)
        if bi is None or te is None:
            raise SystemExit(
                f"سرصفحهٔ قابل‌استفاده پیدا نشد (نیاز به ستون‌های «کد متن» و «متن»). "
                f"ردیف {header_row}: {headers!r}"
            )

        patterns: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            cells = list(row)
            while cells and cells[-1] is None:
                cells.pop()
            if not cells:
                continue
            body_id = _parse_body_id(cells[bi] if bi < len(cells) else None)
            if body_id is None or body_id <= 0:
                continue
            title = _cell_str(cells[ti] if ti is not None and ti < len(cells) else "")
            template = _cell_str(cells[te] if te < len(cells) else "")
            vars_ordered = _extract_variables_ordered(template)
            patterns.append(
                {
                    "bodyId": body_id,
                    "title": title,
                    "templateText": template,
                    "variablePlaceholders": vars_ordered,
                    "variableCount": len(vars_ordered),
                }
            )
        patterns.sort(key=lambda x: x["bodyId"])
        return {
            "source": "melipayamak_panel_export_xlsx",
            "xlsxPath": str(path.resolve()),
            "sheet": ws.title,
            "headerRow": header_row,
            "count": len(patterns),
            "patterns": patterns,
        }
    finally:
        wb.close()


def main() -> int:
    ap = argparse.ArgumentParser(description="Import Melipayamak patterns from xlsx → metadata JSON.")
    ap.add_argument("xlsx", type=Path, help="مسیر فایل اکسل صادرشده از پنل")
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT, help="مسیر خروجی JSON")
    args = ap.parse_args()
    src = args.xlsx.expanduser().resolve()
    if not src.is_file():
        print(f"فایل پیدا نشد: {src}", file=sys.stderr)
        return 2
    payload = import_xlsx(src)
    out = args.out.expanduser().resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"نوشت {payload['count']} پترن در {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
