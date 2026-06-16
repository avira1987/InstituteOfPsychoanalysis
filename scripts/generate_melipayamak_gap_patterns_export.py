#!/usr/bin/env python3
"""از روی metadata/sms_pattern_coverage_matrix.json کمبود پترن‌ها را با توضیح فارسی می‌سازد؛

خروجی:
  - metadata/sms_pattern_gaps_descriptions_fa.json
  - metadata/melipayamak_patterns_to_create.xlsx  (برای ثبت دستی در پنل ملی‌پیامک)

اجرای مجدد پس از به‌روزرسانی ماتریس پوشش:
  python scripts/generate_melipayamak_gap_patterns_export.py
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
MATRIX_PATH = ROOT / "metadata" / "sms_pattern_coverage_matrix.json"
OUT_JSON = ROOT / "metadata" / "sms_pattern_gaps_descriptions_fa.json"
OUT_XLSX = ROOT / "metadata" / "melipayamak_patterns_to_create.xlsx"
PROCESSES_DIR = ROOT / "metadata" / "processes"

if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def _process_name_fa(code: str, cache: dict[str, str]) -> str:
    if code in cache:
        return cache[code]
    p = PROCESSES_DIR / f"{code}.json"
    if not p.is_file():
        cache[code] = code
        return code
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
        name = (data.get("process") or {}).get("name_fa") or code
        cache[code] = str(name).strip() or code
    except (OSError, json.JSONDecodeError):
        cache[code] = code
    return cache[code]


# قطعات کلید تمپلیت → برچسب فارسی کوتاه (برای عنوان؛ ناشناخته حذف می‌شود)
_FRAG_FA: dict[str, str] = {
    "absence": "غیبت",
    "recorded": "ثبت‌شده",
    "academic": "آموزشی",
    "calendar": "تقویم",
    "published": "منتشر شده",
    "approved": "تأیید",
    "upload": "بارگذاری",
    "documents": "مدارک",
    "document": "مدرک",
    "attendance": "حضور و غیاب",
    "site": "سایت",
    "manager": "مسئول",
    "delay": "تأخیر",
    "certificate": "گواهی",
    "ready": "آماده",
    "download": "دریافت",
    "change": "تغییر",
    "rejected": "رد",
    "committee": "کمیته",
    "sla": "SLA",
    "breach": "تخطی از مهلت",
    "comprehensive": "جامع",
    "accepted": "پذیرش",
    "application": "درخواست",
    "scientific": "علمی",
    "supervision": "سوپرویژن",
    "invitation": "دعوت",
    "course": "درس",
    "registration": "ثبت‌نام",
    "definitive": "قطعی",
    "suggestion": "پیشنهاد",
    "debt": "بدهی",
    "settlement": "تسویه",
    "required": "الزامی",
    "education": "آموزش",
    "termination": "خاتمه",
    "notice": "اطلاع‌رسانی",
    "extra": "اضافی",
    "session": "جلسه",
    "sessions": "جلسات",
    "supervisor": "سوپروایزر",
    "therapist": "درمانگر",
    "therapy": "درمان آموزشی",
    "student": "دانشجو",
    "cancelled": "لغو",
    "confirmed": "قطعی شدن",
    "payment": "پرداخت",
    "timeout": "اتمام مهلت",
    "request": "درخواست",
    "forwarded": "ارجاع",
    "file": "پرونده",
    "executive": "اجرایی",
    "intern": "انترن",
    "hours": "ساعت",
    "increase": "افزایش",
    "deferred": "تعویق",
    "interview": "مصاحبه",
    "internship": "کارآموزی",
    "started": "آغاز",
    "scheduled": "زمان‌بندی",
    "details": "جزئیات",
    "reminder": "یادآوری",
    "applicant": "متقاضی",
    "inperson": "حضوری",
    "online": "آنلاین",
    "person": "",
    "leave": "مرخصی آموزشی",
    "live": "زنده",
    "credentials": "ورود به سامانه",
    "lms": "",
    "makeup": "جبرانی",
    "proposed": "پیشنهادی",
    "meeting": "جلسه",
    "non": "",
    "invite": "دعوت",
    "patient": "بیمار",
    "referral": "ارجاع",
    "overdue": "معوق",
    "retry": "تلاش مجدد",
    "quota": "سهمیه",
    "exceeded": "پایان",
    "proof": "مدرک تحصیلی",
    "enrollment": "نام‌نویسی",
    "term": "ترم",
    "introductory": "آشنایی",
    "second": "دوم",
    "semester": "ترم",
    "result": "نتیجه",
    "full": "کامل",
    "single": "تک‌درس",
    "conditional": "مشروط",
    "return": "بازگشت",
    "violator": "",
    "violation": "تخلف",
    "no": "عدم",
    "cancellation": "لغو",
    "counter": "پیشنهاد متقابل",
    "proposal": "پیشنهاد",
    "block": "بلوک",
    "transition": "انتقال",
    "complete": "تکمیل",
    "new": "جدید",
    "interruption": "وقفه",
    "multi": "چندگانه",
    "reduction": "کاهش",
    "terminated": "خاتمه",
    "unannounced": "بدون اعلام قبلی",
    "option": "گزینه",
    "standard": "استاندارد",
    "eligible": "واجد شرایط",
    "ineligible": "غیرمؤهل",
    "restart": "آغاز مجدد",
    "changes": "تغییرات",
    "alternative": "زمان جایگزین",
    "blocked": "مسدود",
    "warning": "هشدار",
    "tuition": "شهریه",
    "accounting": "حسابداری",
    "winter": "زمستان",
    "fall": "پاییز",
    "preparation": "آماده‌سازی",
    "upgrade": "ارتقا",
    "ta": "کمک‌مدرس",
    "blog": "وبلاگ",
    "consultation": "مشاوره",
    "essay": "مقاله",
    "questions": "پرسش‌ها",
    "conceptual": "مفهومی",
    "faculty": "هیئت علمی",
    "assistant": "دستیار",
    "congrats": "تبریک",
    "instructor": "مدرس",
    "auto": "خودکار",
    "track": "رسته",
    "completion": "خاتمه رسته",
    "notify": "اطلاع",
    "early": "زودرس",
    "already": "قبلاً",
    "used": "استفاده شده",
    "conditions": "شرایط",
    "met": "",
    "not": "",
    "success": "موفقیت",
    "condition": "شرط",
    "week9": "هفته نهم",
    "deadline": "مهلت",
    "next": "بعدی",
    "open": "شروع",
    "invoice": "فاکتور",
    "suspended": "تعلیق",
    "declined": "انصراف",
    "alert": "اعلان",
    "2term": "دو ترمی",
    "project": "پروژه",
    "progress": "پیشرفت",
    "monitoring": "نظارت",
    "officer": "مسئول",
    "deputy": "معاون",
    "list": "فهرست",
    "deficiency": "نواقص",
}


def _hint_from_template_key(key: str) -> str:
    parts = []
    for tok in key.split("_"):
        fa = _FRAG_FA.get(tok.lower())
        if fa:
            parts.append(fa)
    if parts:
        return "؛ ".join(parts)
    return key.replace("_", " ")


def _audience_label_fa(audience: str) -> str:
    return {
        "student_applicant": "دانشجو یا متقاضی",
        "staff": "کارمندان (استاد، درمانگر، سوپروایزر، مسئول دفتر، کمیته و …)",
        "mixed": "هم‌زمان دانشجو و کارمند مرتبط",
        "other": "سایر نقش‌ها یا لیست‌های گیرنده",
    }.get(audience, audience)


def _extract_placeholder_keys(sms: str) -> list[str]:
    return list(dict.fromkeys(re.findall(r"\{(\w+)\}", sms)))


def _placeholder_hints(placeholders: list[str]) -> str:
    if not placeholders:
        return "در صورت نیاز متغیر با {0}، {1}، … طبق قوانین پنل تعریف شود."
    return "؛ ".join(f"{{{i}}} ← {{{p}}} (نام متغیر در کد)" for i, p in enumerate(placeholders))


def main() -> int:
    import openpyxl  # pylint: disable=import-outside-toplevel
    from openpyxl.styles import Alignment  # pylint: disable=import-outside-toplevel
    from openpyxl.utils import get_column_letter  # pylint: disable=import-outside-toplevel

    from app.services.notification_service import TEMPLATES  # pylint: disable=import-outside-toplevel

    if not MATRIX_PATH.is_file():
        print(f"Missing {MATRIX_PATH}", file=sys.stderr)
        return 2

    matrix = json.loads(MATRIX_PATH.read_text(encoding="utf-8"))
    gaps = matrix.get("gap_templates_need_new_panel_pattern") or []
    proc_cache: dict[str, str] = {}

    rows_out: list[dict] = []
    for idx, gap in enumerate(gaps, start=1):
        tkey = gap.get("template_key") or ""
        procs = gap.get("process_codes") or []
        roles = gap.get("recipient_roles") or []
        audience = gap.get("audience") or ""
        proc_names = [_process_name_fa(c, proc_cache) for c in procs]
        proc_summary = "، ".join(proc_names[:6])
        if len(proc_names) > 6:
            proc_summary += f" (+{len(proc_names) - 6} فرایند دیگر)"
        roles_str = "، ".join(roles) if roles else "نامشخص"
        tpl_entry = TEMPLATES.get(tkey) if tkey else {}
        sms_body = (tpl_entry or {}).get("sms") if isinstance(tpl_entry, dict) else None
        sms_body = sms_body.strip() if isinstance(sms_body, str) else ""
        ph_keys = _extract_placeholder_keys(sms_body)

        hint = _hint_from_template_key(tkey)
        audience_fa = _audience_label_fa(audience)

        sentence_a = (
            f"در سامانهٔ انستیتو برای کلید پیامکی «{tkey}» هنوز پترن تأییدشده‌ای در ملی‌پیامک "
            f"در نگاشت فعلی در نظر گرفته نشده است."
        )
        sentence_b = (
            f"گیرندگان این پیام طبق فرایند: {audience_fa} ({roles_str}). "
            f"فرایندهای مرتبط شامل «{proc_summary}» است."
        )
        sentence_c = (
            "برای ارسال از خط خدماتی اشتراکی باید متن همسان در پنل ملی‌پیامک به‌صورت پترن با متغیرهای "
            "{0}، {1}، … ثبت و تأیید شود؛ سپس bodyId در برنامه یا تنظیمات نگاشت شود."
        )
        gap_paragraph = " ".join([sentence_a, sentence_b, sentence_c])

        suggested_title = f"{proc_names[0] if proc_names else 'عمومی'} — {hint}"
        if len(suggested_title) > 120:
            suggested_title = suggested_title[:117] + "…"

        rows_out.append(
            {
                "row_index": idx,
                "template_key": tkey,
                "suggested_panel_pattern_title_fa": suggested_title,
                "gap_description_fa": gap_paragraph,
                "audience_code": audience,
                "audience_fa": audience_fa,
                "recipient_roles": roles,
                "process_codes": procs,
                "process_names_fa": proc_names,
                "current_free_text_sms_in_code": sms_body or None,
                "code_placeholder_names": ph_keys,
                "variable_mapping_note_fa": _placeholder_hints(ph_keys),
            }
        )

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(
        json.dumps(
            {
                "source_matrix": str(MATRIX_PATH.relative_to(ROOT))
                if MATRIX_PATH.is_relative_to(ROOT)
                else str(MATRIX_PATH),
                "generated_note_fa": "هر ردیف یک پترن پیشنهادی برای ایجاد در پنل ملی‌پیامک است.",
                "row_count": len(rows_out),
                "gaps": rows_out,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "پترنهای_ضروری"
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass

    headers = [
        "ردیف",
        "کلید تمپلیت در برنامه",
        "عنوان پیشنهادی پترن در پنل ملی‌پیامک",
        "شرح کمبود و سناریو (فارسی)",
        "مخاطب",
        "نقش‌های گیرنده در فرایند",
        "کدهای فرایند",
        "نام فارسی فرایندها",
        "متن فعلی پیامک در کد (آزاد)",
        "نام متغیرهای فعلی در متن کد",
        "یادداشت تبدیل به پترن ({0};{1};…)",
    ]
    ws.append(headers)
    for r in rows_out:
        ws.append(
            [
                r["row_index"],
                r["template_key"],
                r["suggested_panel_pattern_title_fa"],
                r["gap_description_fa"],
                r["audience_fa"],
                "، ".join(r["recipient_roles"]),
                "، ".join(r["process_codes"]),
                "؛ ".join(r["process_names_fa"]),
                r["current_free_text_sms_in_code"] or "",
                "، ".join(r["code_placeholder_names"]),
                r["variable_mapping_note_fa"],
            ]
        )

    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1):
        for cell in row:
            cell.alignment = wrap

    col_widths = [5, 24, 40, 52, 26, 30, 26, 42, 50, 30, 42]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)

    print(f"JSON -> {OUT_JSON} ({len(rows_out)} rows)")
    print(f"XLSX -> {OUT_XLSX}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
