"""خروجی گزارش: CSV، XLSX، PDF (چیدمان استاندارد و جدولی)."""

from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any, Literal

import arabic_reshaper
import jdatetime
from bidi.algorithm import get_display
from fpdf import FPDF
from fpdf.enums import Align, TableBordersLayout, TableCellFillMode, VAlign
from fpdf.fonts import FontFace
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

ReportFormat = Literal["csv", "xlsx", "pdf"]

_ASSETS_DIR = Path(__file__).resolve().parent.parent / "assets" / "fonts"
FONT_PATH = _ASSETS_DIR / "Vazirmatn-Regular.ttf"
FONT_BOLD_PATH = _ASSETS_DIR / "Vazirmatn-Bold.ttf"

# پالت ملایم (خاکستری–آبی کم‌رنگ) برای چاپ کم‌مصرف و خوانایی
_COLOR_TITLE_BAR_BG = (229, 231, 235)
_COLOR_TITLE_TEXT = (55, 65, 81)
_COLOR_SECTION_BG = (243, 244, 246)
_COLOR_SECTION_TEXT = (75, 85, 99)
_COLOR_TABLE_HEADER_BG = (220, 226, 234)
_COLOR_TABLE_HEADER_TEXT = (55, 65, 81)
_COLOR_TABLE_STRIPE = (249, 250, 251)
_COLOR_TABLE_BORDER = (209, 213, 219)

# فونت و چگالی ردیف (کوچک‌تر = صرفه‌جویی در کاغذ)
_PDF_FONT_TITLE = 11
_PDF_FONT_SECTION = 8
_PDF_FONT_BODY = 7
_PDF_FONT_FOOTER = 6
_PDF_LINE_TABLE = 4.8
_PDF_LINE_TABLE_WIDE = 4.2
_PDF_PAD_CELL = 1
_PDF_MARGIN_MM = 10


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return str(round(v, 4))
    return str(v).replace("\r\n", " ").replace("\n", " ").strip()


def _fa_pdf_line(text: str) -> str:
    if not text:
        return ""
    t = str(text).replace("\r", " ").replace("\n", " ")
    try:
        reshaped = arabic_reshaper.reshape(t)
        return get_display(reshaped)
    except Exception:
        return t


def _fa_row(cells: list[Any]) -> list[str]:
    return [_fa_pdf_line(_cell_str(c)) for c in cells]


def _rtl_pdf_columns(cells: list[str]) -> list[str]:
    """fpdf ستون‌ها را چپ‌به‌راست می‌کشد؛ برای RTL ستون اول منطقی باید سمت راست صفحه باشد."""
    if len(cells) <= 1:
        return cells
    return list(reversed(cells))


def _rtl_col_widths(col_widths: float | tuple[float, ...] | None) -> float | tuple[float, ...] | None:
    """هم‌راستا با معکوس‌سازی سلول‌ها؛ عرض هر ستون با همان سلول جابه‌جا می‌شود."""
    if col_widths is None:
        return None
    if isinstance(col_widths, (int, float)):
        return float(col_widths)
    t = tuple(float(x) for x in col_widths)
    if len(t) <= 1:
        return t
    return tuple(reversed(t))


def _fa_row_rtl(raw_cells: list[Any]) -> list[str]:
    """reshape + bidi در هر سلول، سپس معکوس ترتیب ستون‌ها (ستون منطقی اول = سمت راست صفحه)."""
    return _rtl_pdf_columns(_fa_row(raw_cells))


def _blocks_by_empty_rows(rows: list[list[Any]]) -> list[list[list[Any]]]:
    """تقسیم به بلوک‌ها با ردیف خالی."""
    blocks: list[list[list[Any]]] = []
    cur: list[list[Any]] = []
    for r in rows:
        if not r:
            if cur:
                blocks.append(cur)
                cur = []
        else:
            cur.append(r)
    if cur:
        blocks.append(cur)
    return blocks


def _normalize_block(block: list[list[Any]]) -> list[list[str]]:
    if not block:
        return []
    n = max(len(r) for r in block)
    out: list[list[str]] = []
    for r in block:
        cells = [_cell_str(r[i]) if i < len(r) else "" for i in range(n)]
        out.append(cells)
    return out


def _is_single_title_row(block: list[list[Any]]) -> bool:
    """یک ردیف تک‌ستونه = عنوان بخش."""
    return len(block) == 1 and len(block[0]) == 1 and bool(_cell_str(block[0][0]).strip())


def rows_to_csv_bytes(rows: list[list[Any]]) -> bytes:
    """CSV با جداکنندهٔ سمی‌کالن، BOM، پایان خط ویندوز."""
    buf = io.StringIO()
    writer = csv.writer(
        buf,
        delimiter=";",
        quoting=csv.QUOTE_MINIMAL,
        doublequote=True,
        lineterminator="\r\n",
    )
    for row in rows:
        writer.writerow([_cell_str(c) for c in row])
    body = buf.getvalue()
    return "\ufeff".encode("utf-8") + body.encode("utf-8")


def rows_to_xlsx_bytes(rows: list[list[Any]], sheet_title: str = "گزارش") -> bytes:
    """اکسل با RTL."""
    wb = Workbook()
    ws = wb.active
    safe_title = (sheet_title or "گزارش")[:31]
    ws.title = safe_title
    ws.sheet_view.rightToLeft = True
    r = 1
    for row in rows:
        if not row:
            r += 1
            continue
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c_idx, value=_cell_str(val))
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        r += 1
    if ws.max_row == 0:
        ws.cell(row=1, column=1, value="(بدون داده)")
        ws.column_dimensions["A"].width = 24
    else:
        for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            letter = get_column_letter(col[0].column)
            max_len = 10
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, min(len(str(cell.value)), 90))
            ws.column_dimensions[letter].width = min(max_len + 2, 60)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _register_fonts(pdf: FPDF) -> None:
    pdf.add_font("Vazir", "", str(FONT_PATH))
    if FONT_BOLD_PATH.is_file():
        pdf.add_font("Vazir", "B", str(FONT_BOLD_PATH))


def _heading_style() -> FontFace:
    return FontFace(
        family="Vazir",
        emphasis="BOLD",
        size_pt=_PDF_FONT_BODY,
        color=_COLOR_TABLE_HEADER_TEXT,
        fill_color=_COLOR_TABLE_HEADER_BG,
    )


class ReportPDF(FPDF):
    """PDF گزارش با فوتر ثابت (نام گیرنده + زمان شمسی) در پایین هر صفحه."""

    def __init__(self, footer_name: str, footer_ts: str) -> None:
        super().__init__(orientation="L", unit="mm", format="A4")
        self._footer_name = (footer_name or "").strip()
        self._footer_ts = footer_ts

    def footer(self) -> None:
        self.set_y(-10)
        self.set_font("Vazir", "", _PDF_FONT_FOOTER)
        self.set_text_color(120, 120, 120)
        parts: list[str] = []
        if self._footer_name:
            parts.append(self._footer_name)
        if self._footer_ts:
            parts.append(self._footer_ts)
        if not parts:
            return
        line = " — ".join(parts)
        self.cell(0, 5, _fa_pdf_line(line), align="C", new_x="LMARGIN", new_y="NEXT")


def rows_to_pdf_bytes(
    rows: list[list[Any]],
    document_title: str = "گزارش",
    *,
    recipient_display_name: str = "",
) -> bytes:
    if not FONT_PATH.is_file():
        raise FileNotFoundError(
            f"فونت فارسی برای PDF در مسیر زیر نیست: {FONT_PATH}. "
            "فایل Vazirmatn-Regular.ttf را در app/assets/fonts قرار دهید."
        )

    footer_ts = jdatetime.datetime.now().strftime("%Y/%m/%d %H:%M")
    pdf = ReportPDF(recipient_display_name, footer_ts)
    _register_fonts(pdf)
    pdf.set_margins(_PDF_MARGIN_MM, _PDF_MARGIN_MM, _PDF_MARGIN_MM)
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.add_page()

    # ─── هدر عنوان (ملایم) ───
    pdf.set_fill_color(*_COLOR_TITLE_BAR_BG)
    pdf.set_text_color(*_COLOR_TITLE_TEXT)
    pdf.set_draw_color(*_COLOR_TABLE_BORDER)
    pdf.set_line_width(0.15)
    pdf.set_font(
        "Vazir",
        "B" if FONT_BOLD_PATH.is_file() else "",
        _PDF_FONT_TITLE,
    )
    title_line = _fa_pdf_line(document_title)
    pdf.cell(0, 9, title_line, align="C", fill=True, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)
    pdf.set_text_color(30, 30, 30)

    # ─── بدنه: بلوک‌ها (بدون جدول جداگانهٔ گیرنده) ───
    for block in _blocks_by_empty_rows(rows):
        if not block:
            continue
        if _is_single_title_row(block):
            txt = _cell_str(block[0][0])
            pdf.set_fill_color(*_COLOR_SECTION_BG)
            pdf.set_text_color(*_COLOR_SECTION_TEXT)
            pdf.set_font("Vazir", "B" if FONT_BOLD_PATH.is_file() else "", _PDF_FONT_SECTION)
            pdf.cell(0, 6, _fa_pdf_line(txt), fill=True, align="C", new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(30, 30, 30)
            pdf.set_font("Vazir", "", _PDF_FONT_BODY)
            pdf.ln(1)
            continue

        norm = _normalize_block(block)
        if not norm:
            continue
        ncols = len(norm[0])
        if ncols == 0:
            continue

        use_head = len(norm) >= 2
        col_fracs = tuple([1.0] * ncols)
        line_h = _PDF_LINE_TABLE if ncols <= 8 else _PDF_LINE_TABLE_WIDE

        pdf.set_font("Vazir", "", _PDF_FONT_BODY)
        pdf.set_draw_color(*_COLOR_TABLE_BORDER)
        pdf.set_line_width(0.12)

        with pdf.table(
            col_widths=_rtl_col_widths(col_fracs),
            width=pdf.epw,
            first_row_as_headings=False,
            num_heading_rows=0,
            text_align=Align.R,
            v_align=VAlign.M,
            line_height=line_h,
            borders_layout=TableBordersLayout.ALL,
            cell_fill_color=_COLOR_TABLE_STRIPE,
            cell_fill_mode=TableCellFillMode.ROWS,
            padding=_PDF_PAD_CELL,
            gutter_height=0.25,
        ) as table:
            if use_head:
                hs = _heading_style()
                r0 = table.row()
                for t in _fa_row_rtl(norm[0]):
                    r0.cell(t, align=Align.C, style=hs)
                for row in norm[1:]:
                    table.row(_fa_row_rtl(row))
            else:
                for row in norm:
                    table.row(_fa_row_rtl(row))
        pdf.ln(1.5)

    out = pdf.output()
    if isinstance(out, (bytes, bytearray)):
        return bytes(out)
    if isinstance(out, str):
        return out.encode("utf-8")
    return bytes(out)


def export_report(
    rows: list[list[Any]],
    fmt: ReportFormat,
    *,
    document_title: str,
    recipient_display_name: str = "",
) -> tuple[bytes, str, str]:
    f = fmt.lower().strip()
    if f == "csv":
        return rows_to_csv_bytes(rows), "text/csv; charset=utf-8", "csv"
    if f == "xlsx":
        return (
            rows_to_xlsx_bytes(rows, sheet_title=document_title[:31]),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xlsx",
        )
    if f == "pdf":
        return (
            rows_to_pdf_bytes(
                rows,
                document_title=document_title,
                recipient_display_name=recipient_display_name,
            ),
            "application/pdf",
            "pdf",
        )
    raise ValueError(f"فرمت نامعتبر: {fmt}")
